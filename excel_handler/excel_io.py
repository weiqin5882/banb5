from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill



def read_excel(path: Path) -> pd.DataFrame:
    return pd.read_excel(path)



def export_result_to_excel(result_df: pd.DataFrame, summary: dict) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df = result_df.copy()
        export_df = export_df.rename(
            columns={
                "index": "序号",
                "order_id": "订单号",
                "product_name": "商品名称",
                "sales_amount": "销售金额",
                "cost_amount": "成本",
                "profit": "单笔利润",
                "final_status": "状态标记",
            }
        )[["序号", "订单号", "商品名称", "销售金额", "成本", "单笔利润", "状态标记"]]

        export_df.to_excel(writer, index=False, sheet_name="对账结果")
        ws = writer.book["对账结果"]

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")

        for row_idx, is_loss in enumerate(result_df["is_loss"].tolist(), start=2):
            if is_loss:
                for col_idx in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = red_fill
                    cell.font = red_font

        start_row = len(export_df) + 3
        ws.cell(row=start_row, column=1, value="汇总")
        ws.cell(row=start_row, column=2, value=f"总销售额: {summary['total_sales']:.2f}")
        ws.cell(row=start_row, column=3, value=f"总成本: {summary['total_cost']:.2f}")
        ws.cell(row=start_row, column=4, value=f"总利润: {summary['total_profit']:.2f}")
        ws.cell(row=start_row, column=5, value=f"订单总数: {summary['order_count']}")

    return output.getvalue()



def validate_file_extension(filename: str) -> bool:
    lower = filename.lower()
    return lower.endswith((".xlsx", ".xls", ".et"))
